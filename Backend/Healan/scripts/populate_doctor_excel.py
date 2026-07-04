"""Populate Docs/Doctor.xlsx Sheet1 (API catalog + seed data) from Document.xlsx and project requirements."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl import Workbook

def _find_repo_doctor_root() -> Path:
    for parent in Path(__file__).resolve().parents:
        if (parent / "Docs" / "Doctor.xlsx").exists():
            return parent
    raise FileNotFoundError("Docs/Doctor.xlsx not found")


ROOT = _find_repo_doctor_root()
DOCS = ROOT / "Docs"
DOCTOR_PATH = DOCS / "Doctor.xlsx"
DOCUMENT_PATH = DOCS / "Document.xlsx"

MICROSERVICE = "HealanWebApi"

# Medical services for seed — all categories from CategoryTypeId
MEDICAL_SERVICES = [
    ("ویزیت پزشک عمومی", "GV-001", "GeneralVisit", 350_000, "ویزیت و معاینه پزشک عمومی"),
    ("ویزیت مجدد پزشک عمومی", "GV-002", "GeneralVisit", 200_000, "ویزیت مجدد تا ۱۰ روز"),
    ("ویزیت پزشک متخصص", "SV-001", "SpecialistVisit", 550_000, "ویزیت پزشک متخصص"),
    ("ویزیت متخصص قلب", "SV-002", "SpecialistVisit", 600_000, "ویزیت متخصص قلب و عروق"),
    ("ویزیت متخصص اطفال", "SV-003", "SpecialistVisit", 520_000, "ویزیت متخصص اطفال"),
    ("آزمایش CBC", "LAB-001", "Lab", 180_000, "شمارش کامل خون"),
    ("آزمایش قند خون ناشتا", "LAB-002", "Lab", 120_000, "FBS"),
    ("آزمایش چربی خون", "LAB-003", "Lab", 160_000, "LDL/HDL/TG"),
    ("آزمایش TSH", "LAB-004", "Lab", 140_000, "عملکرد تیروئید"),
    ("آزمایش ادرار", "LAB-005", "Lab", 90_000, "Urinalysis"),
    ("رادیوگرافی قفسه سینه", "IMG-001", "Imaging", 250_000, "CXR"),
    ("سونوگرافی شکم", "IMG-002", "Imaging", 420_000, "سونوگرافی کامل شکم"),
    ("سونوگرافی تیروئید", "IMG-003", "Imaging", 380_000, "سونوگرافی تیروئید"),
    ("نوار قلب", "IMG-004", "Imaging", 200_000, "ECG"),
    ("بخیه زخم", "PRC-001", "Procedure", 350_000, "بخیه ساده"),
    ("تزریق عضلانی", "NUR-001", "Nursing", 80_000, "تزریق IM"),
    ("سرم‌تراپی وریدی", "NUR-002", "Nursing", 150_000, "IV"),
    ("پانسمان", "NUR-003", "Nursing", 100_000, "پانسمان زخم"),
    ("فیزیوتراپی", "THR-001", "Therapy", 200_000, "هر جلسه فیزیوتراپی"),
    ("مشاوره تخصصی", "CON-001", "Consultation", 300_000, "مشاوره حضوری"),
    ("مشاوره تغذیه", "CON-002", "Consultation", 280_000, "مشاوره تغذیه"),
]

# Insurance companies from Doctor.xlsx Sheet2 (پزشک section)
INSURANCE_COMPANIES = [
    ("تامین اجتماعی", "TAMIN", "Primary"),
    ("سلامت", "SALAMAT", "Primary"),
    ("نیرو مسلح", "NIROO", "Primary"),
    ("آزاد", "FREE", "None"),
]

EXTRA_APIS = [
    (MICROSERVICE, "Dashboard", "Stats", "آمار داشبورد کلینیک", "Table"),
    (MICROSERVICE, "Appointment", "CurrentAppointmentList", "لیست نوبت‌های جاری / صف انتظار", "Table"),
    (MICROSERVICE, "Appointment", "ChangeStatus", "تغییر وضعیت نوبت", None),
    (MICROSERVICE, "Appointment", "GetPaymentMethodTypes", "انواع روش پرداخت", "enum"),
    (MICROSERVICE, "Signature", "ValidateCertificate", "اعتبارسنجی گواهی امضا", None),
    (MICROSERVICE, "Signature", "PdfDigestForMultiSign", "امضای چندگانه PDF", None),
    (MICROSERVICE, "Cardboard", "UserCardboard", "کارتابل کاربر", "WorkFlow"),
    (MICROSERVICE, "User", "CurrentUser", "کاربر جاری", None),
]


def load_document_apis() -> list[tuple]:
    if not DOCUMENT_PATH.exists():
        return []
    wb = openpyxl.load_workbook(DOCUMENT_PATH, data_only=True)
    ws = wb.active
    rows: list[tuple] = []
    current_module: str | None = None
    current_controller: str | None = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        module, fa_desc, controller, action, table, kind, _ = (list(row) + [None] * 7)[:7]
        if module:
            current_module = str(module).strip()
        if controller:
            current_controller = str(controller).strip()

        ctrl = current_controller
        if not ctrl or not action:
            continue

        desc = (fa_desc or current_module or ctrl or "").strip()
        addition = table or kind or ""
        rows.append((
            MICROSERVICE,
            ctrl,
            str(action).strip(),
            desc,
            str(addition).strip() if addition else None,
        ))
    wb.close()
    return rows


def build_sheet1_rows() -> list[tuple]:
    rows: list[tuple] = []
    seen: set[tuple] = set()

    def add(r: tuple):
        if r[2] == "Seed":
            key = (r[1], r[2], r[3])
        else:
            key = (r[1], r[2])
        if key in seen:
            return
        seen.add(key)
        rows.append(r)

    for r in load_document_apis():
        add(r)
    for r in EXTRA_APIS:
        add(r)

    for title, code, category, price, desc in MEDICAL_SERVICES:
        add((
            MICROSERVICE,
            "ServiceType",
            "Seed",
            title,
            f"Seed;Code={code};Category={category};Price={price};Description={desc}",
        ))

    for name, code, ins_type in INSURANCE_COMPANIES:
        add((
            MICROSERVICE,
            "InsuranceCompany",
            "Seed",
            name,
            f"Seed;Code={code};InsuranceType={ins_type}",
        ))

    return rows


def populate():
    wb = openpyxl.load_workbook(DOCTOR_PATH)
    ws1 = wb["Sheet1"]
    ws1.delete_rows(1, ws1.max_row)

    ws1.append(["MicroserviceName", "Controller", "Action", "Description", "AdditionData"])
    for row in build_sheet1_rows():
        ws1.append(list(row))

    wb.save(DOCTOR_PATH)
    print(f"Updated {DOCTOR_PATH} — Sheet1 rows: {ws1.max_row - 1}")


if __name__ == "__main__":
    populate()
